from openpyxl import load_workbook
from openpyxl.styles import Border, Side

def format_excel_with_borders():
    input_file = "Lab5.xlsx"
    output_file = "Lab5_With_Borders.xlsx"
    
    # Load the original workbook
    wb = load_workbook(input_file)
    ws = wb.active
    
    # Define border styles
    thin = Side(style="thin", color="000000")
    thick = Side(style="thick", color="000000")
    
    # Create borders
    general_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    thick_right_border = Border(left=thin, right=thick, top=thin, bottom=thin)
    
    # Remove column F (index 6 in 1-based)
    ws.delete_cols(6)
    print("✅ Removed column F")
    
    # Apply GENERAL BORDER around A4:Z15 (now A4:Y15 after removing F)
    for row in range(4, 16):  # Rows 4-15
        for col in range(1, 26):  # Columns A-Y (1-25)
            ws.cell(row, col).border = general_border
    
    # Apply THICK RIGHT BORDERS to specific columns
    # After removing F: A=1, B=2, C=3, E=5, G=6
    thick_right_columns = [1, 2, 3, 5, 6]  # A, B, C, E, G
    
    for col in thick_right_columns:
        for row in range(4, 16):  # Rows 4-15
            cell = ws.cell(row, col)
            # Apply thick right border while keeping other borders
            cell.border = Border(
                left=cell.border.left or thin,
                right=thick,  # This is the thick right border
                top=cell.border.top or thin,
                bottom=cell.border.bottom or thin
            )
    
    # Save the result
    wb.save(output_file)
    
    print(f"✅ Success! Output saved to: {output_file}")
    print("📋 Applied formatting:")
    print("   - General border around A4:Y15")
    print("   - Thick right border on columns: A, B, C, E, G")
    print("   - Column F removed")
    print("   - All other content and formatting preserved")

# Run the function
format_excel_with_borders()